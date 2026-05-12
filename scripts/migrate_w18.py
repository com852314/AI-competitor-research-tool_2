"""
migrate_w18.py - 一次性遷移腳本
從 archive/W18_V2.xlsx 抽出資料、轉換為 JSON-first 架構

產出：
- data/weekly/W1_2026-04-30.json   （週快照）
- data/master/games.json           （遊戲主表）
- data/master/vendors.json         （廠家對照）
- data/master/platforms.json       （平台對照）

執行方式：
  python migrate_w18.py
（從 scripts/ 資料夾執行）
"""
import json
import os
from pathlib import Path
from openpyxl import load_workbook

# === 路徑 ===
SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
SRC_XLSX = PROJECT_DIR / "archive" / "W18_V2.xlsx"
DATA_DIR = PROJECT_DIR / "data"
WEEKLY_DIR = DATA_DIR / "weekly"
MASTER_DIR = DATA_DIR / "master"

WEEK_NUMBER = 1
CAPTURE_DATE = "2026-04-30"

# === Helper ===
def cell(ws, row, col):
    """取儲存格值，None 轉空字串"""
    v = ws.cell(row=row, column=col).value
    return v if v is not None else ""

def safe_str(v):
    return str(v).strip() if v is not None else ""

def write_json(path, obj):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
    print(f"  ✓ {path.relative_to(PROJECT_DIR)}")

# === 主程式 ===
def main():
    print(f"讀取來源: {SRC_XLSX.relative_to(PROJECT_DIR)}")
    wb = load_workbook(SRC_XLSX, data_only=False)

    # =================================================================
    # 1. 解析「平台對照表」 → platforms.json
    # =================================================================
    print("\n[1/4] 解析平台對照表...")
    ws = wb["平台對照表"]
    # 欄位：平台代號, 平台全名, 市場, 網址, 抓取方式, 是否需登入, 上次抓取, 狀態, 備註
    platforms = []
    for r in range(5, 20):
        code = safe_str(cell(ws, r, 1))
        if not code:
            break
        platforms.append({
            "code": code,
            "fullName": safe_str(cell(ws, r, 2)),
            "market": safe_str(cell(ws, r, 3)),
            "url": safe_str(cell(ws, r, 4)),
            "captureMethod": safe_str(cell(ws, r, 5)),
            "needsLogin": safe_str(cell(ws, r, 6)) == "是",
            "lastCapture": safe_str(cell(ws, r, 7)) or CAPTURE_DATE,
            "status": safe_str(cell(ws, r, 8)) or "active",
            "note": safe_str(cell(ws, r, 9)),
        })
    print(f"  找到 {len(platforms)} 個平台")

    # =================================================================
    # 2. 解析「廠家對照表」 → vendors.json
    # =================================================================
    print("\n[2/4] 解析廠家對照表...")
    ws = wb["廠家對照表"]
    # 欄位：正規化, PT 顯示, BP 顯示, CP 顯示, TCG URL代碼, TCG 顯示, BAJI URL代碼, BAJI 顯示, 備註
    vendors = []
    for r in range(5, 50):
        norm = safe_str(cell(ws, r, 1))
        if not norm:
            break
        v = {
            "normalizedName": norm,
            "displayPerPlatform": {
                "PT": safe_str(cell(ws, r, 2)) or None,
                "BP": safe_str(cell(ws, r, 3)) or None,
                "CP": safe_str(cell(ws, r, 4)) or None,
                "TCG": None,
                "BAJI": None,
            },
            "note": safe_str(cell(ws, r, 9)),
        }
        # TCG: { urlCode, display }
        tcg_url = safe_str(cell(ws, r, 5))
        tcg_disp = safe_str(cell(ws, r, 6))
        if tcg_url or tcg_disp:
            v["displayPerPlatform"]["TCG"] = {
                "urlCode": tcg_url or None,
                "display": tcg_disp or None,
            }
        # BAJI: { urlCode, display }
        baji_url = safe_str(cell(ws, r, 7))
        baji_disp = safe_str(cell(ws, r, 8))
        if baji_url or baji_disp:
            v["displayPerPlatform"]["BAJI"] = {
                "urlCode": baji_url or None,
                "display": baji_disp or None,
            }
        vendors.append(v)
    print(f"  找到 {len(vendors)} 個廠家")

    # =================================================================
    # 3. 解析「遊戲主表」 → games.json
    # =================================================================
    print("\n[3/4] 解析遊戲主表...")
    ws = wb["遊戲主表"]
    # 欄位：主表ID, 正規化名稱, 正規化廠家, PT 名稱, BP 名稱, CP 名稱, TCG ID/名稱, BAJI ID/名稱, 首次發現日期, 備註
    games = []
    for r in range(5, 100):
        gid = cell(ws, r, 1)
        if not gid:
            continue  # 允許跳號
        try:
            gid = int(gid)
        except (ValueError, TypeError):
            continue
        # 解析 TCG / BAJI 欄位（可能含 ID 與名稱）
        def parse_alias(s):
            s = safe_str(s)
            if not s:
                return None
            return s  # 保留原文，未來再規格化

        games.append({
            "masterId": gid,
            "normalizedName": safe_str(cell(ws, r, 2)),
            "normalizedVendor": safe_str(cell(ws, r, 3)),
            "platformAliases": {
                "PT": parse_alias(cell(ws, r, 4)),
                "BP": parse_alias(cell(ws, r, 5)),
                "CP": parse_alias(cell(ws, r, 6)),
                "TCG": parse_alias(cell(ws, r, 7)),
                "BAJI": parse_alias(cell(ws, r, 8)),
            },
            "firstSeen": safe_str(cell(ws, r, 9)) or CAPTURE_DATE,
            "note": safe_str(cell(ws, r, 10)),
        })
    print(f"  找到 {len(games)} 筆遊戲")

    # =================================================================
    # 4. 解析 5 平台的熱門榜 + TCG/BAJI 新遊戲 → W1 weekly snapshot
    # =================================================================
    print("\n[4/4] 解析週快照...")

    def parse_ranking(sheet_name):
        ws = wb[sheet_name]
        rows = []
        for r in range(5, 25):
            rank = cell(ws, r, 3)
            if not rank:
                continue
            try:
                rank = int(rank)
            except (ValueError, TypeError):
                continue
            mid = cell(ws, r, 10)
            try:
                mid = int(mid) if mid not in ("", None) else None
            except (ValueError, TypeError):
                mid = None
            rows.append({
                "rank": rank,
                "name": safe_str(cell(ws, r, 4)),
                "vendor": safe_str(cell(ws, r, 5)),
                "rawVendor": safe_str(cell(ws, r, 6)),
                "masterId": mid,
                "isNewToTop20": safe_str(cell(ws, r, 7)) == "Y",
                "lastWeekRank": cell(ws, r, 8) if cell(ws, r, 8) != "" else None,
                "rankChange": cell(ws, r, 9) if cell(ws, r, 9) != "" else None,
                "note": safe_str(cell(ws, r, 11)),
            })
        return rows

    def parse_newgames(sheet_name):
        ws = wb[sheet_name]
        rows = []
        for r in range(5, 30):
            order = cell(ws, r, 3)
            if not order:
                continue
            try:
                order = int(order)
            except (ValueError, TypeError):
                continue
            rows.append({
                "order": order,
                "name": safe_str(cell(ws, r, 4)),
                "vendor": safe_str(cell(ws, r, 5)),
                "rawVendor": safe_str(cell(ws, r, 6)),
                "gameId": safe_str(cell(ws, r, 7)),
                "isFirstSeen": safe_str(cell(ws, r, 8)) == "Y",
                "note": safe_str(cell(ws, r, 9)),
            })
        return rows

    # 平台對應市場（從上面剛抓的 platforms 推回）
    market_map = {p["code"]: p["market"] for p in platforms}

    snapshot = {
        "week": WEEK_NUMBER,
        "captureDate": CAPTURE_DATE,
        "platforms": {
            "PT": {
                "market": market_map.get("PT", "菲律賓"),
                "rankings": parse_ranking("PT_熱門榜"),
                "newGames": None,
            },
            "BP": {
                "market": market_map.get("BP", "菲律賓"),
                "rankings": parse_ranking("BP_熱門榜"),
                "newGames": None,
            },
            "CP": {
                "market": market_map.get("CP", "菲律賓"),
                "rankings": parse_ranking("CP_熱門榜"),
                "newGames": None,
            },
            "TCG": {
                "market": market_map.get("TCG", "菲律賓"),
                "rankings": parse_ranking("TCG_熱門榜"),
                "newGames": parse_newgames("TCG_新遊戲"),
            },
            "BAJI": {
                "market": market_map.get("BAJI", "孟加拉"),
                "rankings": parse_ranking("BAJI_熱門榜"),
                "newGames": parse_newgames("BAJI_新遊戲"),
            },
        },
    }
    total_rankings = sum(len(snapshot["platforms"][p]["rankings"]) for p in snapshot["platforms"])
    total_newgames = sum(
        len(snapshot["platforms"][p]["newGames"] or [])
        for p in snapshot["platforms"]
    )
    print(f"  Top 20 共 {total_rankings} 筆 / 新遊戲共 {total_newgames} 筆")

    # =================================================================
    # 寫檔
    # =================================================================
    print("\n寫出 JSON：")
    write_json(MASTER_DIR / "platforms.json", {"platforms": platforms})
    write_json(MASTER_DIR / "vendors.json", {"vendors": vendors})
    write_json(MASTER_DIR / "games.json", {"games": games})
    write_json(WEEKLY_DIR / f"W{WEEK_NUMBER}_{CAPTURE_DATE}.json", snapshot)

    print("\n✓ 遷移完成")

if __name__ == "__main__":
    main()
